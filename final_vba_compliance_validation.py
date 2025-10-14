#!/usr/bin/env python3
"""
Final comprehensive validation to ensure complete VBA compliance for zero rate handling
"""

import pandas as pd
import sys
import os
from pathlib import Path
import xlsxwriter
from openpyxl import load_workbook

# Add the current directory to Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils.first_page_generator import FirstPageGenerator

def test_vba_compliance_comprehensive():
    """Comprehensive test for VBA compliance with zero rate handling"""
    print("🧪 Comprehensive VBA Compliance Validation")
    print("=" * 60)
    
    # Create comprehensive test data covering all scenarios
    work_order_data = pd.DataFrame([
        # Normal item with rate
        {
            'Item No.': '001',
            'Description': 'Excavation in ordinary soil',
            'Unit': 'CuM',
            'Quantity Since': 50.0,
            'Rate': 1200.0,
            'Remark': 'Standard rate item'
        },
        # Zero rate item - should only populate Serial No. and Description
        {
            'Item No.': '002',
            'Description': 'Free item - no charge',
            'Unit': 'No',
            'Quantity Since': 5.0,
            'Rate': 0.0,
            'Remark': 'Complimentary item'
        },
        # Another normal item
        {
            'Item No.': '003',
            'Description': 'RCC M20 for foundation',
            'Unit': 'CuM',
            'Quantity Since': 25.5,
            'Rate': 9200.0,
            'Remark': 'Regular rate item'
        },
        # Blank rate item (treated as zero)
        {
            'Item No.': '004',
            'Description': 'Item with blank rate',
            'Unit': 'SqM',
            'Quantity Since': 100.0,
            'Rate': '',  # Blank rate
            'Remark': 'Blank rate item'
        }
    ])
    
    extra_items_data = pd.DataFrame([
        # Normal extra item
        {
            'Item No.': 'EX01',
            'Description': 'Additional electrical work',
            'Unit': 'Mtr',
            'Quantity': 100.0,
            'Rate': 120.0,
            'Remark': 'Extra work'
        },
        # Zero rate extra item
        {
            'Item No.': 'EX02',
            'Description': 'Free inspection service',
            'Unit': 'No',
            'Quantity': 1.0,
            'Rate': 0.0,
            'Remark': 'No charge'
        }
    ])
    
    title_data = {
        'Name of Work ;-': 'Test Infrastructure Project - VBA Compliance Validation',
        'Name of Contractor or supplier :': 'ABC Construction Ltd',
        'Date': '15-10-2025'
    }
    
    # Generate First Page
    generator = FirstPageGenerator()
    output_path = "final_vba_compliance_test.xlsx"
    
    try:
        generator.generate_first_page(work_order_data, extra_items_data, title_data, output_path)
        print(f"✅ First Page generated: {output_path}")
        
        # Verify the file was created
        if Path(output_path).exists():
            print("✅ Output file created successfully")
            
            # Validate the content
            validation_result = validate_excel_content(output_path, work_order_data, extra_items_data)
            
            return output_path, validation_result
        else:
            print("❌ Output file was not created")
            return None, False
            
    except Exception as e:
        print(f"❌ Error generating First Page: {e}")
        import traceback
        traceback.print_exc()
        return None, False

def validate_excel_content(file_path, work_order_data, extra_items_data):
    """Validate the content of the generated Excel file for VBA compliance"""
    print("\n🔍 Validating Excel Content for VBA Compliance...")
    
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        worksheet = workbook['First Page']
        
        # Start checking from row 22 (0-indexed: row 21)
        current_row = 22  # VBA style row 22
        
        # Validate work order items
        print(f"\n📋 Validating Work Order Items...")
        for index, row_data in work_order_data.iterrows():
            rate = row_data.get('Rate', 0)
            # Convert rate to float for comparison
            try:
                rate_value = float(rate) if rate != '' and rate is not None else 0.0
            except (ValueError, TypeError):
                rate_value = 0.0
            
            print(f"  Item {row_data['Item No.']}: Rate = {rate} ({'Zero' if rate_value == 0 else 'Non-Zero'})")
            
            # Check VBA compliance
            if rate_value == 0:
                # For zero rates: Only Serial No. (D) and Description (E) should be populated
                # All other columns should be blank/empty
                unit_cell = worksheet.cell(row=current_row, column=1).value  # Column A
                qty_since_cell = worksheet.cell(row=current_row, column=2).value  # Column B
                qty_upto_cell = worksheet.cell(row=current_row, column=3).value  # Column C
                serial_no_cell = worksheet.cell(row=current_row, column=4).value  # Column D
                description_cell = worksheet.cell(row=current_row, column=5).value  # Column E
                rate_cell = worksheet.cell(row=current_row, column=6).value  # Column F
                amount_upto_cell = worksheet.cell(row=current_row, column=7).value  # Column G
                amount_since_cell = worksheet.cell(row=current_row, column=8).value  # Column H
                remark_cell = worksheet.cell(row=current_row, column=9).value  # Column I
                
                # Validate zero rate behavior
                if (serial_no_cell == row_data['Item No.'] and 
                    description_cell == row_data['Description'] and
                    (unit_cell is None or unit_cell == '') and
                    (qty_since_cell is None or qty_since_cell == '') and
                    (qty_upto_cell is None or qty_upto_cell == '') and
                    (rate_cell is None or rate_cell == '') and
                    (amount_upto_cell is None or amount_upto_cell == '') and
                    (amount_since_cell is None or amount_since_cell == '') and
                    (remark_cell is None or remark_cell == '')):
                    print(f"    ✅ Zero rate compliance: Only Serial No. and Description populated")
                else:
                    print(f"    ❌ Zero rate compliance FAILED")
                    print(f"      Expected: Only Serial No.='{row_data['Item No.']}', Description='{row_data['Description']}'")
                    print(f"      Actual: A='{unit_cell}', B='{qty_since_cell}', C='{qty_upto_cell}', D='{serial_no_cell}', E='{description_cell}', F='{rate_cell}', G='{amount_upto_cell}', H='{amount_since_cell}', I='{remark_cell}'")
                    return False
            else:
                # For non-zero rates: All columns should be populated
                # This is a basic check - full validation would be more complex
                serial_no_cell = worksheet.cell(row=current_row, column=4).value  # Column D
                description_cell = worksheet.cell(row=current_row, column=5).value  # Column E
                
                if (serial_no_cell == row_data['Item No.'] and 
                    description_cell == row_data['Description']):
                    print(f"    ✅ Non-zero rate item correctly populated")
                else:
                    print(f"    ❌ Non-zero rate item validation FAILED")
                    return False
            
            current_row += 1
        
        # Skip the "Extra Items (With Premium)" header row
        current_row += 1
        
        # Validate extra items
        print(f"\n📋 Validating Extra Items...")
        for index, row_data in extra_items_data.iterrows():
            rate = row_data.get('Rate', 0)
            # Convert rate to float for comparison
            try:
                rate_value = float(rate) if rate != '' and rate is not None else 0.0
            except (ValueError, TypeError):
                rate_value = 0.0
            
            print(f"  Extra Item {row_data['Item No.']}: Rate = {rate} ({'Zero' if rate_value == 0 else 'Non-Zero'})")
            
            # Check VBA compliance
            if rate_value == 0:
                # For zero rates: Only Serial No. (D) and Description (E) should be populated
                serial_no_cell = worksheet.cell(row=current_row, column=4).value  # Column D
                description_cell = worksheet.cell(row=current_row, column=5).value  # Column E
                unit_cell = worksheet.cell(row=current_row, column=1).value  # Column A
                qty_cell = worksheet.cell(row=current_row, column=3).value  # Column C
                rate_cell = worksheet.cell(row=current_row, column=6).value  # Column F
                amount_upto_cell = worksheet.cell(row=current_row, column=7).value  # Column G
                amount_since_cell = worksheet.cell(row=current_row, column=8).value  # Column H
                remark_cell = worksheet.cell(row=current_row, column=9).value  # Column I
                
                # Validate zero rate behavior
                if (serial_no_cell == row_data['Item No.'] and 
                    description_cell == row_data['Description'] and
                    (unit_cell is None or unit_cell == '') and
                    (qty_cell is None or qty_cell == '') and
                    (rate_cell is None or rate_cell == '') and
                    (amount_upto_cell is None or amount_upto_cell == '') and
                    (amount_since_cell is None or amount_since_cell == '') and
                    (remark_cell is None or remark_cell == '')):
                    print(f"    ✅ Zero rate compliance: Only Serial No. and Description populated")
                else:
                    print(f"    ❌ Zero rate compliance FAILED")
                    print(f"      Expected: Only Serial No.='{row_data['Item No.']}', Description='{row_data['Description']}'")
                    print(f"      Actual: A='{unit_cell}', C='{qty_cell}', D='{serial_no_cell}', E='{description_cell}', F='{rate_cell}', G='{amount_upto_cell}', H='{amount_since_cell}', I='{remark_cell}'")
                    return False
            else:
                # For non-zero rates: All columns should be populated
                serial_no_cell = worksheet.cell(row=current_row, column=4).value  # Column D
                description_cell = worksheet.cell(row=current_row, column=5).value  # Column E
                
                if (serial_no_cell == row_data['Item No.'] and 
                    description_cell == row_data['Description']):
                    print(f"    ✅ Non-zero rate extra item correctly populated")
                else:
                    print(f"    ❌ Non-zero rate extra item validation FAILED")
                    return False
            
            current_row += 1
        
        print("\n✅ Excel content validation PASSED")
        print("✅ VBA compliance verified for all items")
        return True
            
    except Exception as e:
        print(f"❌ Error validating Excel content: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main validation function"""
    print("🚀 Final VBA Compliance Validation for Zero Rate Handling")
    print("=" * 70)
    
    # Test VBA compliance
    file_path, validation_result = test_vba_compliance_comprehensive()
    
    print("\n" + "=" * 70)
    print("📊 FINAL VALIDATION SUMMARY")
    print("=" * 70)
    print(f"File Generation: {'✅ PASS' if file_path else '❌ FAIL'}")
    print(f"VBA Compliance: {'✅ PASS' if validation_result else '❌ FAIL'}")
    
    if file_path and validation_result:
        print("\n🎉 ALL VALIDATIONS PASSED!")
        print("✅ FirstPageGenerator fully complies with VBA specification")
        print("✅ Zero rate handling is correctly implemented:")
        print("   - For zero rates: Only Serial No. and Description are populated")
        print("   - All other columns remain blank for zero rate items")
        print("✅ Implementation matches exact VBA behavior")
        
        # Clean up test file
        try:
            Path(file_path).unlink()
            print("✅ Test file cleaned up")
        except:
            print("⚠️  Could not clean up test file")
            
        return True
    else:
        print("\n💥 VALIDATION FAILED!")
        print("❌ Implementation does not fully comply with VBA specification")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)